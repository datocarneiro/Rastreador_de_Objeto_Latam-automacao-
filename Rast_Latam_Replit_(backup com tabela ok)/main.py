import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from flask import Flask, render_template, request, session
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException

# Configuração do aplicativo Flask
app = Flask(__name__)
app.secret_key = 'dato123'  # Defina uma chave secreta adequada

##########################################################################
'''
# para rodar no replit usar essas configuraçõa
options = Options()
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
options.headless = False  # Executar o Chrome de forma oculta

driver = webdriver.Chrome(options=options)
'''
#############################################################################

# para rodar local usar essa configuração aqui
from webdriver_manager.chrome import ChromeDriverManager

servico = Service(ChromeDriverManager().install())

opcoes = Options()
opcoes.headless = True  # modo off ou não
driver = webdriver.Chrome(service=servico, options=opcoes)

# Variável global para armazenar a lista de pendentes
lista_pendentes = []

@app.route('/')
def index():
    return render_template('index.html', pendentes=lista_pendentes)

@app.route('/resultado')
def resultado():
    statuses, datas, df = capturar_status_pendentes()
    table_html = df.to_html(classes='table table-bordered', index=False)
    return render_template('resultado.html', table_html=table_html)

@app.route('/', methods=['POST'])
def preparar_dados_planilha():
    global lista_pendentes  # Acessando a variável global
    file = request.files['file']

    if not file.filename.endswith('.xlsx'):
        return "Por favor, selecione um arquivo Excel (.xlsx)"

    # Salvar o arquivo Excel em uma pasta temporária
    filename = secure_filename(file.filename)
    file.save(filename)
    session['excel_filename'] = filename  # Armazenar o nome do arquivo na sessão

    planilha = load_workbook(filename)
    aba_ativa = planilha.active

    lista_pendentes = []
    for coluna_a, coluna_c, coluna_d in zip(aba_ativa["A"][1:],
                                            aba_ativa["C"][1:],
                                            aba_ativa["D"][1:]):
        if coluna_d.value != 'ENTREGUE':
            if coluna_c.value is not None:
                lista_pendentes.append(coluna_c.value)

    print("=" * 150)
    print(f'As pendente de entrega são: {lista_pendentes}')
    print("=" * 150)

    return render_template('index.html', pendentes=lista_pendentes)

def captura_status(awb):
    driver.get(
        f"https://www.latamcargo.com/en/trackshipment?docNumber={awb}&docPrefix=957&soType=SO"
    )

    wait = WebDriverWait(driver, 30)

    try:
        status_evento = wait.until(
            EC.visibility_of_element_located(
                (By.XPATH, '//*[@id="statusTable"]/tbody/tr[1]/td[1]')))
        status = status_evento.text

        data_evento = wait.until(
            EC.visibility_of_element_located(
                (By.XPATH, '//*[@id="statusTable"]/tbody/tr[1]/td[6]')))
        data = data_evento.text

        captura = f'dados capturados,AWB,{awb},STATUS,{status},DATA_EVENTO,{data}'
        print(captura)

        return status, data
    except TimeoutException:
        status = "Erro de tempo limite"
        data = "Erro de tempo limite"
        print("Erro de tempo limite ao capturar os dados")
        return status, data

def capturar_status_pendentes():
    dados_rastreamento = []
    statuses = []
    datas = []

    # Obter o nome do arquivo Excel da sessão
    excel_filename = session.get('excel_filename')

    if excel_filename:
        planilha = load_workbook(excel_filename)
        aba_ativa = planilha.active

        for coluna_a, coluna_c, coluna_d in zip(aba_ativa["A"][1:],
                                                aba_ativa["C"][1:],
                                                aba_ativa["D"][1:]):
            if coluna_d.value != 'ENTREGUE':
                if coluna_c.value is not None:
                    awb = coluna_c.value
                    status, data = captura_status(awb)
                    dados_rastreamento.append({
                        'AWB': awb,
                        'STATUS': status,
                        'DATA_EVENTO': data
                    })
                    statuses.append(status)
                    datas.append(data)

        df = pd.DataFrame(dados_rastreamento)

        return statuses, datas, df
    else:
        return [], [], None  # Retorna vazios se o arquivo não estiver definido na sessão

if __name__ == '__main__':
    app.run(host="0.0.0.0", port=8080)
